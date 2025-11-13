import streamlit as st
import pandas as pd
from openpyxl import load_workbook

# -------------------------
# File paths
# -------------------------
EXCEL_FILE = "C:/Users/jjete/OneDrive/1. Life Stuff/Budget/Bank Data Export/budget_export.xlsm"
SHEET_NAME = "Quicklook"

# -------------------------
# Load Quicklook with styles
# -------------------------
@st.cache_data
def load_quicklook_with_styles():
    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb[SHEET_NAME]

    def extract_range(start_row, end_row, start_col, end_col):
        values = []
        colors = []
        for r in range(start_row, end_row + 1):
            row_values = []
            row_colors = []
            for c in range(start_col, end_col + 1):
                cell = ws.cell(row=r, column=c)
                row_values.append(cell.value)
                fill = cell.fill.start_color.rgb
                row_colors.append(f"#{fill[-6:]}" if fill else "#FFFFFF")
            values.append(row_values)
            colors.append(row_colors)
        df = pd.DataFrame(values)
        df_colors = pd.DataFrame(colors)
        return df, df_colors

    df_top, df_top_colors = extract_range(1, 4, 1, 6)
    df_middle, df_middle_colors = extract_range(5, 18, 1, 4)
    df_bottom, df_bottom_colors = extract_range(19, 19, 1, 3)
    e18_value = ws.cell(row=18, column=5).value

    return (df_top, df_top_colors,
            df_middle, df_middle_colors,
            df_bottom, df_bottom_colors,
            e18_value)

# -------------------------
# Combine sections
# -------------------------
def combine_quicklook(df_top, df_middle, df_bottom):
    df_middle_padded = pd.concat([df_middle, pd.DataFrame("", index=df_middle.index, columns=range(4,6))], axis=1)
    df_bottom_padded = pd.concat([df_bottom, pd.DataFrame("", index=df_bottom.index, columns=range(3,6))], axis=1)
    df_combined = pd.concat([df_top, df_middle_padded, df_bottom_padded], ignore_index=True)
    return df_combined

def combine_colors(df_top_colors, df_middle_colors, df_bottom_colors):
    df_middle_colors_padded = pd.concat([df_middle_colors, pd.DataFrame("#FFFFFF", index=df_middle_colors.index, columns=range(4,6))], axis=1)
    df_bottom_colors_padded = pd.concat([df_bottom_colors, pd.DataFrame("#FFFFFF", index=df_bottom_colors.index, columns=range(3,6))], axis=1)
    df_combined_colors = pd.concat([df_top_colors, df_middle_colors_padded, df_bottom_colors_padded], ignore_index=True)
    return df_combined_colors

# -------------------------
# Hide repeated values to simulate merged cells
# -------------------------
def hide_repeats(df):
    df_hide = df.copy()
    for col in df_hide.columns:
        previous = None
        for i in range(len(df_hide)):
            if df_hide.iloc[i, col] == previous:
                df_hide.iloc[i, col] = ""
            else:
                previous = df_hide.iloc[i, col]
    return df_hide

# -------------------------
# Style DataFrame with colors, bold headers, center numbers
# -------------------------
def style_dataframe(df, color_df):
    df_hide = hide_repeats(df)

    def align_cells(val):
        if isinstance(val, (int, float)):
            return "text-align: center;"
        return "text-align: left;"

    # Bold the first row of each section
    def bold_rows(row_index):
        if row_index in [0, 4, 18]:  # top section, middle section start, bottom row
            return ["font-weight: bold;"] * len(df_hide.columns)
        return [""] * len(df_hide.columns)

    styled = (df_hide.style
              .apply(lambda x: color_df.iloc[x.name], axis=1)
              .applymap(align_cells)
              .apply(lambda x: bold_rows(x.name), axis=1)
              .set_table_styles([
                  {'selector': 'tbody tr:nth-child(4n+4)', 'props': [('border-bottom', '2px solid #000')]}  # optional separator
              ])
             )
    return styled

# -------------------------
# Load data
# -------------------------
(df_top, df_top_colors,
 df_middle, df_middle_colors,
 df_bottom, df_bottom_colors,
 e18_value) = load_quicklook_with_styles()

df_combined = combine_quicklook(df_top, df_middle, df_bottom)
color_combined = combine_colors(df_top_colors, df_middle_colors, df_bottom_colors)

# -------------------------
# Streamlit display
# -------------------------
st.set_page_config(page_title="📊 Quicklook Budget Summary", layout="wide")
st.title("📊 Quicklook Budget Summary")

st.dataframe(style_dataframe(df_combined, color_combined), use_container_width=True)

st.subheader("E18 Value")
st.write(e18_value)
